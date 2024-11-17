import pandas as pd
import pywhatkit
import time
import threading
import tkinter as tk
from openpyxl import load_workbook

# Caminho para a planilha
caminho_planilha = "lista.xlsx"

# Carregar a planilha
try:
    df = pd.read_excel(caminho_planilha)
except FileNotFoundError:
    print(f"Erro: A planilha '{caminho_planilha}' não foi encontrada.")
    df = pd.DataFrame(columns=["nomes", "whatsapp", "Mensagem", "Status"])  # Cria uma nova estrutura

# Variáveis globais para controle
paused = False
sending = False  # Indica se o envio está em andamento
thread = None    # Armazena a thread de envio
qtd_erros = 0
qtd_certos = 0

# Função para alternar pausa/retomar
def alternar_pausa():
    global paused
    paused = not paused
    if paused:
        btn_pausar.config(text="Retomar")
    else:
        btn_pausar.config(text="Pausar")


# Função para atualizar o status na planilha
def atualizar_status(caminho, linha, status):
    try:
        workbook = load_workbook(caminho)
        sheet = workbook.active
        sheet.cell(row=linha + 2, column=5, value=status)  # Coluna "Status"
        workbook.save(caminho)
        print(f"Status '{status}' atualizado na linha {linha + 2}.")
    except Exception as e:
        print(f"Erro ao atualizar o status: {e}")


# Função para enviar mensagens no WhatsApp
def enviarmsg():
    global qtd_erros
    global qtd_certos
    global sending, paused
    sending = True
    for index, row in df.iterrows():
        if not sending:  # Interrompe o envio se o usuário parar o processo
            break

        while paused:  # Aguarda enquanto estiver pausado
            time.sleep(1)

        nome = row['Nomes']
        numero = str(row['Whatsapp'])
        numero2 = str(row['Whatsapp2'])
        mensagem = row['Mensagem']
        mensagempt = f"Olá {nome}, {mensagem}"
        numerowhats = f"+{numero}"
        numerowhats2 = f"+{numero2}"
        atuallb.config(text=nome)
        try:
            # Enviar mensagem pelo WhatsApp
            print(f"Enviando mensagem para {nome} no número {numerowhats}...")
            pywhatkit.sendwhatmsg_instantly(
                phone_no=numerowhats,
                message=mensagempt,
                wait_time=5,
                tab_close=True
            )
           # time.sleep(5)
            
            atualizar_status(caminho_planilha, index, "Mensagem enviada")
            qtd_certos += 1           
            qtd_certoslb.config(text=f"{qtd_certos} Enviados")
        except Exception as e:
            print(f"Erro ao enviar mensagem para {nome} ({numerowhats}): {e}")
            print(f"Enviando mensagem para {nome} no número {numerowhats}...")
            pywhatkit.core.core.close_tab()
            try:
                pywhatkit.sendwhatmsg_instantly(
                    phone_no=numerowhats2,
                    message=mensagempt,
                    wait_time=5,
                    tab_close=True
                )
                atualizar_status(caminho_planilha, index, "Mensagem enviada")
                qtd_certos += 1           
                qtd_certoslb.config(text=f"{qtd_certos} Enviados")
            except Exception as e:
                atualizar_status(caminho_planilha, index, "Erro")
                time.sleep(1)  # Tempo para evitar travamentos
                qtd_erros = qtd_erros + 1
                qtd_erroslb.config(text=f"{qtd_erros} Erros")
                pywhatkit.core.core.close_tab()
    sending = False
    print("Envio concluído!")


# Função para iniciar o envio em uma thread separada
def iniciar_envio():
    global thread, sending
    if sending:  # Impede iniciar múltiplos envios ao mesmo tempo
        print("O envio já está em andamento!")
        return

    thread = threading.Thread(target=enviarmsg)  # Cria a thread
    thread.daemon = True  # Permite que a thread seja encerrada com o programa
    thread.start()  # Inicia o envio


# Função para interromper o envio
def parar_envio():
    global sending
    sending = False
    print("Envio interrompido pelo usuário.")


# Interface gráfica com Tkinter
root = tk.Tk()
root.title("Automação WhatsApp")

# Configuração da janela
root.geometry("230x200")
root.attributes('-topmost', True)          # Sempre no topo
root.attributes('-alpha', 0.92)  
# Fundo da aplicação
frame_Fundo = tk.Frame(root, background="#4F4F4F", borderwidth=8, relief="raised")
frame_Fundo.grid(row=0, column=0, sticky="nsew", padx=5)

# Configurar o redimensionamento
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# Botão para iniciar o envio
btn_iniciar = tk.Button(frame_Fundo, text="Iniciar",relief="raised",borderwidth=3, command=iniciar_envio)
btn_iniciar.grid(row=2, column=0, sticky="ew",padx=5,pady=5)

# Botão para pausar/retomar
btn_pausar = tk.Button(frame_Fundo, text="Pausar",relief="raised",borderwidth=3, command=alternar_pausa)
btn_pausar.grid(row=2, column=1, sticky="ew",padx=5,pady=5)

# Botão para parar o envio
btn_parar = tk.Button(frame_Fundo, text="Parar",relief="raised",borderwidth=3, command=parar_envio)
btn_parar.grid(row=2, column=2, sticky="ew",padx=5,pady=5)

atuallb = tk.Label(frame_Fundo,font=("Arial",14),bg='#4F4F4F',fg='white' )
atuallb.grid(row=4,column=0,columnspan=3,padx=5,pady=5)

qtd_erroslb = tk.Label(frame_Fundo,font=("Arial",14),bg='#4F4F4F',fg='red')
qtd_erroslb.grid(row=6,column=0,sticky='nsew',columnspan=3,padx=5,pady=5)

qtd_certoslb = tk.Label(frame_Fundo,font=("Arial",14),bg='#4F4F4F',fg='green')
qtd_certoslb.grid(row=5,column=0,sticky='nsew',columnspan=3,padx=5,pady=5)
# Inicia o loop principal
frame_Fundo.grid_columnconfigure(0, weight=1)
frame_Fundo.grid_columnconfigure(1, weight=1)
frame_Fundo.grid_columnconfigure(2, weight=1)
root.mainloop()
